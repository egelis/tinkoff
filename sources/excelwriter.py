import os
from decimal import Decimal
import datetime

import openpyxl
from openpyxl.styles import Alignment, Font


class ExcelWriter:
    """Класс, записывающий данные портфолио в файл Excel"""

    def __init__(self, filename, sheet, positions, balance, courses, operations):
        self.filename = f'../{filename}.xlsx'
        self.positions = positions
        self.balance = balance
        self.courses = courses
        self.operations = operations

        # Открытие файла и страницы в файле Excel
        # Если файла не существует, то он создается
        if os.path.exists(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook[sheet]
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = sheet
            self.workbook.save(self.filename)

    @staticmethod
    def _make_cell_title(cell, position):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal=f'{position}', vertical="center")

    def write_names_of_columns(self):
        names = ('Название', 'Вид бумаги', 'Котировка', 'Цена, шт.', 'Кол-во', 'Общая цена, руб.', '%',
                 'Идеал. соотнош.')

        ws = self.worksheet

        # Данным методом идем по строкам (cell - очередная ячейка строки), тут 1 строка
        for cells in self.worksheet['A6':'H6']:
            for i, cell in enumerate(cells):
                self._make_cell_title(cell, position='center')
                cell.value = names[i]

        ws.merge_cells('I6:K6')
        ws['I6'] = 'Правки(%, ₽, $)'
        self._make_cell_title(ws['I6'], position='center')

        self.workbook.save(self.filename)

    def write_courses(self):
        ws = self.worksheet

        for i in range(1, 3):
            self._make_cell_title(ws[f'D{i}'], 'right')

        ws['D1'] = 'Курс доллара:'
        ws['D2'] = 'Курс евро:'
        ws['E1'] = self.courses['USD']
        ws['E2'] = self.courses['EUR']

        self.workbook.save(self.filename)

    def write_balance(self):
        ws = self.worksheet

        self._make_cell_title(ws['A2'], 'right')
        ws['A2'] = 'Баланс:'
        ws['B2'] = self.balance[0].balance  # RUB

        self.workbook.save(self.filename)

    def write_portfolio_price(self):
        ws = self.worksheet

        self._make_cell_title(ws['A1'], 'right')
        ws['A1'] = 'Общая цена портфеля:'
        ws['B1'] = get_portfolio_price(self.balance, self.positions, self.courses)

        self.workbook.save(self.filename)

    def write_ratios(self):
        ws = self.worksheet

        names = ('Акции:', 'Облигации:', 'Золото:', 'Валюта:')
        for i, name in enumerate(names, start=1):
            self._make_cell_title(ws[f'G{i}'], 'right')
            ws[f'G{i}'] = name

        self.workbook.save(self.filename)

    def write_positions(self):
        # Располагаем облигации первыми в списке
        self.positions.sort(key=lambda pos: get_unit_type(pos) != 'Облигации')

        # Данным методом идем по строкам (cell - очередная ячейка строки), тут len(self.positions) строк
        for cells, position in zip(self.worksheet['A7':f'F{len(self.positions) + 7 - 1}'], self.positions):
            unit_price = get_unit_price(position)
            elements_of_position = (f'{position.name}, {position.average_position_price.currency}',
                                    get_unit_type(position),
                                    position.ticker,
                                    unit_price,
                                    position.balance,
                                    get_total_position_price_rub(position, self.courses))

            for i, cell in enumerate(cells, start=0):
                location = 'right' if i in (3, 5) else 'left' if i in (0,) else 'center'
                cell.alignment = Alignment(horizontal=location, vertical="center")
                cell.value = elements_of_position[i]

        self.workbook.save(self.filename)

    def write_positions_percentages(self):
        ws = self.worksheet

        for i in range(7, len(self.positions) + 7):
            ws[f'G{i}'].alignment = Alignment(horizontal='center', vertical="center")
            ws[f'G{i}'] = f'=F{i} / B$1'

        self.workbook.save(self.filename)

    def write_revisions(self):
        ws = self.worksheet

        number_of_bounds = 0
        for position in self.positions:
            if get_unit_type(position) == 'Облигации':
                number_of_bounds += 1

        # Объединение и обобщенный подсчет правок для всех Облигаций
        ws.merge_cells(f'I7:I{6 + number_of_bounds}')
        ws['I7'].alignment = Alignment(horizontal='center', vertical="center")
        ws['I7'] = f'=H7-SUM(G7:G{6 + number_of_bounds})'

        ws.merge_cells(f'J7:J{6 + number_of_bounds}')
        ws['J7'].alignment = Alignment(horizontal='right', vertical="center")
        ws['J7'] = '=I7 * B$1'

        ws.merge_cells(f'K7:K{6 + number_of_bounds}')
        ws['K7'].alignment = Alignment(horizontal='right', vertical="center")
        ws['K7'] = '=J7 / E$1'

        # %
        for i in range(7 + number_of_bounds, len(self.positions) + 7):
            ws[f'I{i}'].alignment = Alignment(horizontal='center', vertical="center")
            ws[f'I{i}'] = f'=H{i} - G{i}'

        # руб.
        for i in range(7 + number_of_bounds, len(self.positions) + 7):
            ws[f'J{i}'].alignment = Alignment(horizontal='right', vertical="center")
            ws[f'J{i}'] = f'=I{i} * B$1'

        # долл.
        for i in range(7 + number_of_bounds, len(self.positions) + 7):
            ws[f'K{i}'].alignment = Alignment(horizontal='right', vertical="center")
            ws[f'K{i}'] = f'=J{i} / E$1'

        self.workbook.save(self.filename)

    def write_pay_in(self):
        print(f'Пополнения: {round(get_sum_pay_in(self.operations), 2)}')
        portfolio_price = get_portfolio_price(self.balance, self.positions, self.courses)
        print(f'Прибыль от вложенной суммы: {round(100 * (1 - get_sum_pay_in(self.operations) / portfolio_price), 2)} %')

    def write_table_to_excel(self):
        self.write_portfolio_price()
        self.write_balance()
        self.write_courses()
        self.write_ratios()

        self.write_names_of_columns()
        self.write_positions()
        self.write_positions_percentages()
        self.write_revisions()

        self.write_pay_in()


def get_unit_type(position) -> str:
    bonds = ('FinEx Еврооблигации рос. компаний (RUB)',
             'FinEx Еврооблигации рос. компаний (USD)',
             'FinEx Казначейские облигации США (USD)',
             'FinEx Казначейские облигации США')

    gold = ('FinEx Золото',)

    currency = ('Доллар США',
                'Евро')

    if position.instrument_type == 'Bond' or position.name in bonds:
        return 'Облигации'
    elif position.name in currency:
        return 'Валюта'
    elif position.name in gold:
        return 'Золото'
    else:
        return 'Акции'


def get_unit_price(position) -> Decimal:
    return Decimal(round(position.average_position_price.value + (position.expected_yield.value /
                                                                  position.balance), 2))


def get_total_position_price_rub(position, courses) -> Decimal:
    currency = position.average_position_price.currency.value

    quantity = position.balance
    purchase_price = position.average_position_price.value
    expected_yield = position.expected_yield.value

    if currency == 'RUB':
        return Decimal(quantity * purchase_price + expected_yield)
    elif currency == 'USD':
        return Decimal((quantity * purchase_price + expected_yield) * courses['USD'])
    elif currency == 'EUR':
        return Decimal((quantity * purchase_price + expected_yield) * courses['EUR'])


def get_portfolio_price(balance, positions, courses) -> Decimal:
    price = Decimal(0)
    for position in positions:
        price += get_total_position_price_rub(position, courses)
    price += Decimal(balance[0].balance)  # Баланс в RUB

    return price


def get_sum_pay_in(operations) -> Decimal:
    sum_pay_in = Decimal(0)
    get_candle_from_date = operations[1]
    for operation in operations[0]:
        if operation.operation_type.value == "PayIn" or operation.operation_type.value == "PayOut":
            if operation.currency.value == "USD":
                course_from_date = get_candle_from_date(
                    "BBG0013HGFT4",
                    str(operation.date - datetime.timedelta(minutes=15)).replace(" ", "T"),
                    str(operation.date).replace(" ", "T")
                ).payload.candles[0].c

                sum_pay_in += Decimal(str(operation.payment * course_from_date))
            else:
                sum_pay_in += Decimal(str(operation.payment))
    return sum_pay_in
